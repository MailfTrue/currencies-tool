import csv
import io
from typing import Iterable
from collections import defaultdict

import openpyxl
import openpyxl.utils
import openpyxl.worksheet.dimensions
import openpyxl.styles
import pdfkit

from jinja import env as jinja_env


def render_csv(header: Iterable, rows: Iterable[Iterable]) -> io.StringIO:
    csv_file = io.StringIO()
    csv_writer = csv.writer(csv_file)
    csv_writer.writerow(header)
    csv_writer.writerows(rows)
    csv_file.seek(0)
    return csv_file


def render_xlsx(header: Iterable, rows: Iterable[Iterable]) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active

    header = list(header)
    column_sizes = defaultdict(lambda: 20)
    column_sizes.update({k: len(v) for (k, v) in enumerate(header, 1)})

    ws.append(header)
    if 'header' not in wb.named_styles:
        header_style = openpyxl.styles.NamedStyle(name="header")
        header_style.font = openpyxl.styles.Font(color='495057', bold=True)
        header_style.alignment = openpyxl.styles.Alignment(horizontal="center")
        wb.add_named_style(header_style)
    for cell in ws["1"]:
        cell.style = "header"

    for row in rows:
        row = list(row)
        for i, col in enumerate(row, 1):
            length = len(str(col))
            if length > column_sizes[i]:
                column_sizes[i] = length
        ws.append(row)

    dim_holder = openpyxl.worksheet.dimensions.DimensionHolder(worksheet=ws)
    for col in range(ws.min_column, ws.max_column + 1):
        col_dim = openpyxl.worksheet.dimensions.ColumnDimension(ws, min=col, max=col, width=column_sizes[col] + 5)
        dim_holder[openpyxl.utils.get_column_letter(col)] = col_dim
    ws.column_dimensions = dim_holder

    xlsx_file = io.BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file


def render_jinja(template: str, *args, **kwargs) -> str:
    return jinja_env.get_template(template).render(*args, **kwargs)


def render_pdf(html: str) -> io.BytesIO:
    return io.BytesIO(pdfkit.from_string(html))
